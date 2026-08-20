"""ClientImportService – CSV-/Excel-Massenimport fuer die Mandanten-
datenbank (20.08.).

Zwei Schritte, bewusst getrennt:
1. `parse_csv`/`parse_xlsx`: liest die Rohdatei in eine Liste von
   Zeilen-Dicts mit NORMALISIERTEN Spaltennamen (`_HEADER_ALIASES`
   uebersetzt gaengige deutsche/englische Kopfzeilen-Varianten auf die
   internen Feldnamen) - wirft `ImportFileError`, wenn die Datei gar
   nicht lesbar ist oder die beiden Pflichtspalten (Name/Mandantennummer)
   im Header komplett fehlen (dann kann keine einzige Zeile sinnvoll
   verarbeitet werden).
2. `import_clients`: verarbeitet die geparsten Zeilen EINZELN ueber
   `ClientService.create_client` (dieselbe Pflichtfeld-/Eindeutigkeits-
   pruefung wie die manuelle Erfassung, siehe app/clients/service.py) -
   eine fehlerhafte Zeile (leeres Pflichtfeld, doppelte Mandantennummer)
   wird uebersprungen und mit Zeilennummer+Grund gesammelt, bricht aber
   NICHT den gesamten Import ab (sonst wuerde eine einzige Tippfehler-
   Zeile einen 500-Zeilen-Import komplett verhindern).
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.clients.service import ClientValidationError, create_client

# Interner Feldname -> akzeptierte Kopfzeilen-Varianten (kleingeschrieben,
# Leerzeichen/Bindestriche/Unterstriche werden vor dem Abgleich entfernt,
# siehe _normalize_header). Reine UI-/Import-Bequemlichkeit, keine
# vollstaendige Aufzaehlung jeder denkbaren Schreibweise.
_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("name", "mandant", "mandantname", "kanzleiname", "kunde"),
    "client_number": (
        "mandantennummer",
        "mandantennr",
        "clientnumber",
        "kundennummer",
        "nummer",
    ),
    "contact_email": ("email", "emailadresse", "mail"),
    "contact_phone": ("telefon", "phone", "tel", "telefonnummer"),
    "practice_area": ("rechtsgebiet", "practicearea", "fachgebiet"),
}
_REQUIRED_FIELDS = ("name", "client_number")


class ImportFileError(Exception):
    """Datei nicht lesbar oder Pflichtspalten (Name/Mandantennummer) im
    Header nicht vorhanden - der gesamte Import wird abgebrochen, BEVOR
    irgendeine Zeile verarbeitet wird."""


@dataclass
class ImportRowError:
    row_number: int  # 1-basiert, zaehlt die Datenzeile (ohne Header)
    reason: str


@dataclass
class ImportResult:
    created_count: int = 0
    errors: list[ImportRowError] = field(default_factory=list)

    @property
    def has_errors(self) -> bool:
        return len(self.errors) > 0


def _normalize_header(raw: str) -> str:
    return raw.strip().lower().replace(" ", "").replace("-", "").replace("_", "")


def _build_header_index_map(headers: list[str]) -> dict[str, int]:
    """Liefert internes Feld -> Spaltenindex fuer alle ERKANNTEN Spalten
    (unbekannte Spalten werden stillschweigend ignoriert - z. B. eine
    zusaetzliche "Notizen"-Spalte in einer Alt-Excel-Datei)."""
    normalized = [_normalize_header(h) for h in headers]
    field_to_index: dict[str, int] = {}
    for field_name, aliases in _HEADER_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                field_to_index[field_name] = index
                break
    return field_to_index


def _require_field_columns_present(field_to_index: dict[str, int]) -> None:
    missing = [f for f in _REQUIRED_FIELDS if f not in field_to_index]
    if missing:
        readable = {"name": "Name", "client_number": "Mandantennummer"}
        missing_readable = ", ".join(readable[m] for m in missing)
        raise ImportFileError(
            f"Pflichtspalte(n) nicht gefunden: {missing_readable}. "
            "Erwartete Kopfzeile z. B. 'Name' und 'Mandantennummer'."
        )


def _row_to_dict(row_values: list[str], field_to_index: dict[str, int]) -> dict[str, str]:
    result: dict[str, str] = {}
    for field_name, index in field_to_index.items():
        if index < len(row_values):
            result[field_name] = (row_values[index] or "").strip()
    return result


def parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("cp1252")
        except UnicodeDecodeError as exc:
            raise ImportFileError(
                "CSV-Datei konnte nicht gelesen werden (Zeichenkodierung) - "
                "bitte als UTF-8 oder Windows-1252 speichern."
            ) from exc

    # Trennzeichen-Erkennung: deutsche Excel-Exporte nutzen haeufig ";"
    # statt ",". `csv.Sniffer` ist zuverlaessiger als ein fester Wert.
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        raise ImportFileError("CSV-Datei ist leer.")

    field_to_index = _build_header_index_map(rows[0])
    _require_field_columns_present(field_to_index)
    return [_row_to_dict(row, field_to_index) for row in rows[1:]]


def parse_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl wirft diverse Fehlertypen
        raise ImportFileError(
            "Excel-Datei konnte nicht gelesen werden (beschaedigt oder kein "
            "gueltiges .xlsx-Format)."
        ) from exc

    sheet = workbook.active
    rows_raw = list(sheet.iter_rows(values_only=True))
    workbook.close()
    rows = [
        [str(cell).strip() if cell is not None else "" for cell in row]
        for row in rows_raw
        if any(cell is not None and str(cell).strip() for cell in row)
    ]
    if not rows:
        raise ImportFileError("Excel-Datei ist leer.")

    field_to_index = _build_header_index_map(rows[0])
    _require_field_columns_present(field_to_index)
    return [_row_to_dict(row, field_to_index) for row in rows[1:]]


def import_clients(db: Session, rows: list[dict[str, str]], *, actor: str) -> ImportResult:
    """Verarbeitet bereits geparste Zeilen (siehe Moduldocstring) - jede
    Zeile wird EINZELN committet (`create_client(commit=True)`, Default).
    Bewusst NICHT ein einziger Sammel-Commit am Ende: `db.rollback()` nach
    einer fehlgeschlagenen Zeile wuerde sonst auch bereits erfolgreich
    verarbeitete Zeilen derselben Transaktion mit zuruecknehmen (SQLAlchemy-
    Rollback wirkt auf die GESAMTE Session, nicht nur den letzten Flush) -
    ein einzelner Tippfehler in Zeile 400 duerfte die Zeilen 1-399 nicht
    rueckwirkend mit verwerfen."""
    result = ImportResult()
    for offset, row in enumerate(rows):
        row_number = offset + 1
        try:
            create_client(
                db,
                name=row.get("name", ""),
                client_number=row.get("client_number", ""),
                contact_email=row.get("contact_email"),
                contact_phone=row.get("contact_phone"),
                practice_area=row.get("practice_area"),
                responsible_user_id=None,
                actor=actor,
            )
            result.created_count += 1
        except ClientValidationError as exc:
            db.rollback()
            result.errors.append(ImportRowError(row_number=row_number, reason=str(exc)))

    return result
